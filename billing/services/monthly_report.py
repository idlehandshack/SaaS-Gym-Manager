"""
billing/services/monthly_report.py
-------------------------------------
Generates per-gym monthly revenue report data and Excel export.
Used exclusively by owner-only views — never called from staff/member views.

Public API:
    get_monthly_report_data(gym, year, month) -> dict
    generate_monthly_report_excel(gym, year, month) -> BytesIO
"""
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth

from billing.models import Invoice, Payment
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Palette (mirrors gst_report.py) ───────────────────────────────────────────
HEADER_FILL  = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
TOTALS_FILL  = PatternFill(fill_type='solid', fgColor='D6E4F0')
TOTALS_FONT  = Font(bold=True, size=10, name='Arial')
TITLE_FONT   = Font(bold=True, size=14, name='Arial', color='1F4E79')
META_FONT    = Font(bold=True, size=10, name='Arial')
META_VAL     = Font(size=10, name='Arial')
DATA_FONT    = Font(size=10, name='Arial')
FOOTER_FONT  = Font(italic=True, size=9, name='Arial', color='808080')
THIN         = Side(style='thin')
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INR_FMT      = '[\u20b9-4009]#,##0.00'
CENTER       = Alignment(horizontal='center', vertical='center')


# ── Core data function ─────────────────────────────────────────────────────────

def get_monthly_report_data(gym, year: int, month: int) -> dict:
    """
    Returns a dict with all KPIs for the given gym/year/month.
    All queries are scoped to gym= so cross-gym leakage is impossible.

    Keys returned:
        year, month, month_label,
        total_invoices, total_paid, total_pending, total_grand,
        total_members,
        invoices        — list of Invoice objects (issued, ordered by invoice_number)
        daily_breakdown — list of {day, paid, pending, invoice_count}
        plan_breakdown  — list of {plan_name, count, paid, pending}
        method_breakdown — list of {method_label, count, paid}
    """
    start = date(year, month, 1)
    # Last day of month
    if month == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month + 1, 1)

    # ── Invoices (from billing) ───────────────────────────────────────────
    invoice_qs = (
        Invoice.objects
        .filter(gym=gym, invoice_date__gte=start, invoice_date__lt=end, status='issued')
        .select_related('member', 'member__selectPlan', 'related_payment')
        .order_by('invoice_number')
    )
    invoices = list(invoice_qs)

    total_invoices = len(invoices)
    total_grand    = sum(inv.grand_total for inv in invoices)
    total_taxable  = sum(inv.taxable_value for inv in invoices)
    total_cgst     = sum(inv.cgst_amount for inv in invoices)
    total_sgst     = sum(inv.sgst_amount for inv in invoices)
    total_igst     = sum(inv.igst_amount for inv in invoices)

    # ── Payments (for paid/pending split) ─────────────────────────────────
    payment_qs = (
        Payment.objects
        .filter(gym=gym, payment_date__gte=start, payment_date__lt=end)
    )
    agg = payment_qs.aggregate(
        total_paid=Sum('paid_amount'),
        total_pending=Sum('pending_amount'),
    )
    total_paid    = agg['total_paid']    or Decimal('0')
    total_pending = agg['total_pending'] or Decimal('0')

    # ── Total unique members who paid this month ───────────────────────────
    total_members = (
        payment_qs
        .values('enrollment_id')
        .distinct()
        .count()
    )

    # ── Daily breakdown ────────────────────────────────────────────────────
    daily_raw = (
        payment_qs
        .values('payment_date')
        .annotate(
            paid=Sum('paid_amount'),
            pending=Sum('pending_amount'),
            invoice_count=Count('id'),
        )
        .order_by('payment_date')
    )
    daily_breakdown = [
        {
            'day':           r['payment_date'].strftime('%d %b'),
            'paid':          float(r['paid'] or 0),
            'pending':       float(r['pending'] or 0),
            'invoice_count': r['invoice_count'],
        }
        for r in daily_raw
    ]

    # ── Plan breakdown ─────────────────────────────────────────────────────
    plan_raw = (
        payment_qs
        .values('plan_name')
        .annotate(
            count=Count('id'),
            paid=Sum('paid_amount'),
            pending=Sum('pending_amount'),
        )
        .order_by('-paid')
    )
    plan_breakdown = [
        {
            'plan_name': r['plan_name'] or '—',
            'count':     r['count'],
            'paid':      float(r['paid'] or 0),
            'pending':   float(r['pending'] or 0),
        }
        for r in plan_raw
    ]

    # ── Payment method breakdown ───────────────────────────────────────────
    METHOD_LABELS = {'C': 'Cash', 'U': 'UPI', 'B': 'UPI + Cash', None: 'Unknown'}
    method_raw = (
        payment_qs
        .values('payment_method')
        .annotate(count=Count('id'), paid=Sum('paid_amount'))
        .order_by('-paid')
    )
    method_breakdown = [
        {
            'method_label': METHOD_LABELS.get(r['payment_method'], r['payment_method'] or '—'),
            'count':        r['count'],
            'paid':         float(r['paid'] or 0),
        }
        for r in method_raw
    ]

    return {
        'year':             year,
        'month':            month,
        'month_label':      start.strftime('%B %Y'),
        'total_invoices':   total_invoices,
        'total_paid':       float(total_paid),
        'total_pending':    float(total_pending),
        'total_grand':      float(total_grand),
        'total_taxable':    float(total_taxable),
        'total_cgst':       float(total_cgst),
        'total_sgst':       float(total_sgst),
        'total_igst':       float(total_igst),
        'total_members':    total_members,
        'invoices':         invoices,
        'daily_breakdown':  daily_breakdown,
        'plan_breakdown':   plan_breakdown,
        'method_breakdown': method_breakdown,
    }


def get_all_months_summary(gym, num_months: int = 12) -> list:
    """
    Returns a list of {month_label, year, month, total_paid, total_invoices}
    for the last `num_months` months, newest first.
    Used to populate the month-picker on the dashboard.
    """
    from django.utils import timezone
    today = timezone.localdate()
    result = []
    y, m = today.year, today.month
    for _ in range(num_months):
        start = date(y, m, 1)
        end   = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
        agg = (
            Payment.objects
            .filter(gym=gym, payment_date__gte=start, payment_date__lt=end)
            .aggregate(total_paid=Sum('paid_amount'), count=Count('id'))
        )
        result.append({
            'month_label':    start.strftime('%B %Y'),
            'year':           y,
            'month':          m,
            'total_paid':     float(agg['total_paid'] or 0),
            'total_invoices': agg['count'],
        })
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return result


# ── Excel export ───────────────────────────────────────────────────────────────

def generate_monthly_report_excel(gym, year: int, month: int) -> BytesIO:
    """
    Builds a 2-sheet Excel workbook:
      Sheet 1 — Invoice Detail   (one row per invoice)
      Sheet 2 — Summary KPIs     (totals + plan/method breakdown)
    Returns a BytesIO ready to stream.
    """
    data = get_monthly_report_data(gym, year, month)
    wb   = Workbook()

    _build_detail_sheet(wb, gym, data)
    _build_summary_sheet(wb, gym, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sheet helpers ──────────────────────────────────────────────────────────────

def _hdr(sheet, headers, widths):
    row = sheet.max_row + 1
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = sheet.cell(row=row, column=i, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = CENTER
        c.border    = THIN_BORDER
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.row_dimensions[row].height = 28


def _style_row(sheet, row_num, ncols):
    for col in range(1, ncols + 1):
        c = sheet.cell(row=row_num, column=col)
        c.font      = DATA_FONT
        c.border    = THIN_BORDER
        c.alignment = CENTER


def _totals_row(sheet, values, ncols):
    sheet.append([])
    row = sheet.max_row + 1
    sheet.append(values)
    for col in range(1, ncols + 1):
        c = sheet.cell(row=row, column=col)
        c.font      = TOTALS_FONT
        c.fill      = TOTALS_FILL
        c.border    = THIN_BORDER
        c.alignment = CENTER
    return row


def _meta(sheet, rows, ncols):
    """Write label:value meta rows and merge value columns."""
    for label, value in rows:
        r = sheet.max_row + 1
        sheet.append([f'{label} :', value])
        sheet.cell(row=r, column=1).font = META_FONT
        sheet.cell(row=r, column=2).font = META_VAL
        sheet.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)


def _build_detail_sheet(wb, gym, data):
    sheet = wb.active
    sheet.title = 'Invoice Detail'
    ncols = 9

    # Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = sheet.cell(row=1, column=1, value=f"{gym.gym_name} — Monthly Report ({data['month_label']})")
    c.font      = TITLE_FONT
    c.alignment = CENTER
    sheet.row_dimensions[1].height = 28
    sheet.append([])

    _meta(sheet, [
        ('Gym',       gym.gym_name),
        ('Period',    data['month_label']),
        ('Generated', datetime.now().strftime('%d-%b-%Y %I:%M %p')),
    ], ncols)

    sheet.column_dimensions['A'].width = 18
    sheet.append([])

    _hdr(sheet,
         ['Invoice No', 'Date', 'Member Name', 'Phone', 'Plan',
          'Taxable', 'GST', 'Grand Total', 'Status'],
         [20, 14, 26, 14, 20, 14, 12, 14, 12])

    header_row = sheet.max_row
    sheet.freeze_panes = f'A{header_row + 1}'

    cur_cols = [6, 7, 8]  # Taxable, GST, Grand Total (1-based)

    for inv in data['invoices']:
        row_num = sheet.max_row + 1
        gst_total = inv.cgst_amount + inv.sgst_amount + inv.igst_amount
        sheet.append([
            inv.invoice_number,
            inv.invoice_date.strftime('%d-%m-%Y'),
            inv.customer_name,
            inv.customer_phone or '—',
            inv.member.selectPlan.plan if (inv.member and inv.member.selectPlan) else '—',
            float(inv.taxable_value),
            float(gst_total),
            float(inv.grand_total),
            inv.get_status_display(),
        ])
        _style_row(sheet, row_num, ncols)
        for col in cur_cols:
            sheet.cell(row=row_num, column=col).number_format = INR_FMT

    if not data['invoices']:
        r = sheet.max_row + 1
        sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = sheet.cell(row=r, column=1, value='No invoices found for this period.')
        c.font      = FOOTER_FONT
        c.alignment = CENTER

    # Totals
    total_gst = data['total_cgst'] + data['total_sgst'] + data['total_igst']
    tr = _totals_row(sheet, [
        'TOTAL', '', '', '', '',
        data['total_taxable'], total_gst, data['total_grand'], '',
    ], ncols)
    sheet.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
    for col in cur_cols:
        sheet.cell(row=tr, column=col).number_format = INR_FMT

    # Footer
    sheet.append([])
    r = sheet.max_row + 1
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = sheet.cell(row=r, column=1,
                   value='Generated by EnterGYM. Confidential — for gym owner use only.')
    c.font      = FOOTER_FONT
    c.alignment = Alignment(horizontal='left')


def _build_summary_sheet(wb, gym, data):
    sheet = wb.create_sheet('Summary')
    ncols = 4

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = sheet.cell(row=1, column=1, value=f"Summary — {data['month_label']}")
    c.font      = TITLE_FONT
    c.alignment = CENTER
    sheet.row_dimensions[1].height = 28
    sheet.append([])

    _meta(sheet, [
        ('Total Invoices',  data['total_invoices']),
        ('Total Collected', f"₹{data['total_paid']:,.2f}"),
        ('Total Pending',   f"₹{data['total_pending']:,.2f}"),
        ('Grand Total',     f"₹{data['total_grand']:,.2f}"),
        ('Active Members',  data['total_members']),
    ], ncols)

    sheet.column_dimensions['A'].width = 22
    sheet.column_dimensions['B'].width = 30

    # Plan breakdown
    sheet.append([])
    _hdr(sheet, ['Plan Name', 'Members', 'Collected (₹)', 'Pending (₹)'],
         [26, 14, 18, 18])
    for row in data['plan_breakdown']:
        r = sheet.max_row + 1
        sheet.append([row['plan_name'], row['count'], row['paid'], row['pending']])
        _style_row(sheet, r, 4)
        for col in [3, 4]:
            sheet.cell(row=r, column=col).number_format = INR_FMT

    # Method breakdown
    sheet.append([])
    _hdr(sheet, ['Payment Method', 'Count', 'Collected (₹)', ''],
         [26, 14, 18, 18])
    for row in data['method_breakdown']:
        r = sheet.max_row + 1
        sheet.append([row['method_label'], row['count'], row['paid'], ''])
        _style_row(sheet, r, 4)
        sheet.cell(row=r, column=3).number_format = INR_FMT

    # Footer
    sheet.append([])
    r = sheet.max_row + 1
    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = sheet.cell(row=r, column=1,
                   value='Generated by EnterGYM. Confidential — for gym owner use only.')
    c.font      = FOOTER_FONT
    c.alignment = Alignment(horizontal='left')