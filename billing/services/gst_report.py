"""
billing/services/gst_report.py
"""
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from billing.models import Invoice

# ── Palette ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10, name='Arial')
HEADER_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
TOTALS_FILL   = PatternFill(fill_type='solid', fgColor='D6E4F0')
TOTALS_FONT   = Font(bold=True, size=10, name='Arial')
TITLE_FONT    = Font(bold=True, size=14, name='Arial', color='1F4E79')
META_FONT     = Font(bold=True, size=10, name='Arial')
META_VAL_FONT = Font(size=10, name='Arial')
FOOTER_FONT   = Font(italic=True, size=9, name='Arial', color='808080')
DATA_FONT     = Font(size=10, name='Arial')
THIN          = Side(style='thin')
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER        = Alignment(horizontal='center', vertical='center')


def _inr(value) -> str:
    try:
        return f"₹{float(value):,.2f}"
    except (TypeError, ValueError):
        return '₹0.00'


# ── Helpers ────────────────────────────────────────────────────────────────────

def _set_headers(sheet, headers, col_widths, start_col=1):
    row = sheet.max_row + 1
    for i, (header, width) in enumerate(zip(headers, col_widths), start=start_col):
        cell = sheet.cell(row=row, column=i, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER
        sheet.column_dimensions[get_column_letter(i)].width = width
    sheet.row_dimensions[row].height = 30


def _style_data_row(sheet, row_num, num_cols, start_col=1):
    for col in range(start_col, start_col + num_cols):
        cell = sheet.cell(row=row_num, column=col)
        cell.font      = DATA_FONT
        cell.border    = THIN_BORDER
        cell.alignment = CENTER


def _merge_write(sheet, row, start_col, end_col, value, font=None, alignment=None):
    sheet.merge_cells(
        start_row=row, start_column=start_col,
        end_row=row,   end_column=end_col
    )
    cell = sheet.cell(row=row, column=start_col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    return cell


# ── Report builder ─────────────────────────────────────────────────────────────

def generate_gstr1_style_report(gym, start_date, end_date) -> BytesIO:
    invoices = list(
        Invoice.objects
        .filter(gym=gym, invoice_date__range=(start_date, end_date), status='issued')
        .prefetch_related('line_items')
        .order_by('invoice_date', 'invoice_number')
    )

    wb = Workbook()
    _build_sales_sheet(wb, gym, invoices, start_date, end_date)
    _build_summary_sheet(wb, gym, invoices, start_date, end_date)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Sheet 1: GST Sales Report ──────────────────────────────────────────────────

def _build_sales_sheet(wb, gym, invoices, start_date, end_date):
    sheet = wb.active
    sheet.title = 'GST Sales Report'
    num_cols = 11
    generated_on = datetime.now().strftime('%d-%b-%Y %I:%M %p')

    _merge_write(
        sheet, row=1, start_col=1, end_col=num_cols,
        value='EnterGYM GST Sales Report',
        font=TITLE_FONT,
        alignment=CENTER,
    )
    sheet.row_dimensions[1].height = 28
    sheet.append([])

    gst_profile = getattr(gym, 'gst_profile', None)
    meta_rows = [
        ('Gym Name',      gym.gym_name),
        ('GSTIN',         gst_profile.gstin if gst_profile else '—'),
        ('Business Name', gst_profile.legal_business_name if gst_profile else '—'),
        ('Period',        f"{start_date.strftime('%d-%b-%Y')}  to  {end_date.strftime('%d-%b-%Y')}"),
        ('Generated On',  generated_on),
    ]
    for label, value in meta_rows:
        row_num = sheet.max_row + 1
        sheet.append([f'{label} :', value])
        sheet.cell(row=row_num, column=1).font = META_FONT
        sheet.cell(row=row_num, column=2).font = META_VAL_FONT
        sheet.merge_cells(start_row=row_num, start_column=2,
                          end_row=row_num,   end_column=num_cols)

    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 26
    sheet.append([])

    headers    = [
        'Invoice No', 'Date', 'Customer Name', 'Invoice Type',
        'Taxable Value (₹)', 'CGST (₹)', 'SGST (₹)', 'IGST (₹)',
        'Total GST (₹)', 'Grand Total (₹)', 'Status',
    ]
    col_widths = [18, 14, 26, 16, 18, 14, 14, 14, 14, 16, 12]

    table_header_row = sheet.max_row + 1
    _set_headers(sheet, headers, col_widths)
    sheet.freeze_panes = f'A{table_header_row + 1}'

    # Running totals
    tot_taxable = Decimal('0')
    tot_cgst    = Decimal('0')
    tot_sgst    = Decimal('0')
    tot_igst    = Decimal('0')
    tot_gst     = Decimal('0')
    tot_grand   = Decimal('0')

    for inv in invoices:
        total_gst = inv.cgst_amount + inv.sgst_amount + inv.igst_amount
        tot_taxable += inv.taxable_value
        tot_cgst    += inv.cgst_amount
        tot_sgst    += inv.sgst_amount
        tot_igst    += inv.igst_amount
        tot_gst     += total_gst
        tot_grand   += inv.grand_total

        row_num = sheet.max_row + 1
        sheet.append([
            inv.invoice_number,
            inv.invoice_date.strftime('%d-%m-%Y'),
            inv.customer_name,
            inv.get_invoice_type_display(),
            _inr(inv.taxable_value),   # ← string, no number_format
            _inr(inv.cgst_amount),
            _inr(inv.sgst_amount),
            _inr(inv.igst_amount),
            _inr(total_gst),
            _inr(inv.grand_total),
            inv.get_status_display(),
        ])
        _style_data_row(sheet, row_num, num_cols)

    if not invoices:
        empty_row = sheet.max_row + 1
        _merge_write(
            sheet, row=empty_row, start_col=1, end_col=num_cols,
            value='No invoices found for the selected period.',
            font=Font(italic=True, size=10, name='Arial', color='808080'),
            alignment=CENTER,
        )
        sheet.row_dimensions[empty_row].height = 24

    # Totals row
    sheet.append([])
    totals_row = sheet.max_row + 1
    sheet.append([
        'TOTAL', '', '', '',
        _inr(tot_taxable),
        _inr(tot_cgst),
        _inr(tot_sgst),
        _inr(tot_igst),
        _inr(tot_gst),
        _inr(tot_grand),
        '',
    ])
    sheet.merge_cells(start_row=totals_row, start_column=1,
                      end_row=totals_row,   end_column=4)
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=totals_row, column=col)
        cell.font      = TOTALS_FONT
        cell.fill      = TOTALS_FILL
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
    sheet.row_dimensions[totals_row].height = 20

    sheet.append([])
    _merge_write(
        sheet, row=sheet.max_row + 1, start_col=1, end_col=num_cols,
        value='This report was generated automatically by EnterGYM.',
        font=FOOTER_FONT,
        alignment=Alignment(horizontal='left'),
    )
    _merge_write(
        sheet, row=sheet.max_row + 1, start_col=1, end_col=num_cols,
        value='For GST return filing, please consult your Chartered Accountant.',
        font=FOOTER_FONT,
        alignment=Alignment(horizontal='left'),
    )


# ── Sheet 2: GST Summary ──────────────────────────────────────────────────────

def _build_summary_sheet(wb, gym, invoices, start_date, end_date):
    sheet = wb.create_sheet('GST Summary')
    num_cols = 7
    generated_on = datetime.now().strftime('%d-%b-%Y %I:%M %p')

    _merge_write(
        sheet, row=1, start_col=1, end_col=num_cols,
        value='EnterGYM GST Summary',
        font=TITLE_FONT,
        alignment=CENTER,
    )
    sheet.row_dimensions[1].height = 28
    sheet.append([])

    gst_profile = getattr(gym, 'gst_profile', None)
    meta_rows = [
        ('Gym Name',      gym.gym_name),
        ('GSTIN',         gst_profile.gstin if gst_profile else '—'),
        ('Business Name', gst_profile.legal_business_name if gst_profile else '—'),
        ('Period',        f"{start_date.strftime('%d-%b-%Y')}  to  {end_date.strftime('%d-%b-%Y')}"),
        ('Generated On',  generated_on),
    ]
    for label, value in meta_rows:
        row_num = sheet.max_row + 1
        sheet.append([f'{label} :', value])
        sheet.cell(row=row_num, column=1).font = META_FONT
        sheet.cell(row=row_num, column=2).font = META_VAL_FONT
        sheet.merge_cells(start_row=row_num, start_column=2,
                          end_row=row_num,   end_column=num_cols)

    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 26
    sheet.append([])

    rate_groups: dict[Decimal, dict] = defaultdict(lambda: {
        'invoice_ids': set(),
        'taxable': Decimal('0'),
        'cgst':    Decimal('0'),
        'sgst':    Decimal('0'),
        'igst':    Decimal('0'),
    })
    for inv in invoices:
        for item in inv.line_items.all():
            g = rate_groups[item.gst_rate]
            g['invoice_ids'].add(inv.pk)
            g['taxable'] += item.taxable_value
            g['cgst']    += item.cgst_amount
            g['sgst']    += item.sgst_amount
            g['igst']    += item.igst_amount

    headers    = [
        'GST Rate (%)', 'Invoice Count',
        'Taxable Value (₹)', 'CGST (₹)', 'SGST (₹)', 'IGST (₹)', 'Total GST (₹)',
    ]
    col_widths = [16, 16, 20, 14, 14, 14, 16]

    table_header_row = sheet.max_row + 1
    _set_headers(sheet, headers, col_widths)
    sheet.freeze_panes = f'A{table_header_row + 1}'

    if not rate_groups:
        empty_row = sheet.max_row + 1
        _merge_write(
            sheet, row=empty_row, start_col=1, end_col=num_cols,
            value='No invoice data found for the selected period.',
            font=Font(italic=True, size=10, name='Arial', color='808080'),
            alignment=CENTER,
        )
        sheet.row_dimensions[empty_row].height = 24
        return

    grand_inv_ids = set()
    grand_taxable = Decimal('0')
    grand_cgst    = Decimal('0')
    grand_sgst    = Decimal('0')
    grand_igst    = Decimal('0')

    for rate in sorted(rate_groups.keys()):
        g = rate_groups[rate]
        total_gst = g['cgst'] + g['sgst'] + g['igst']

        grand_inv_ids |= g['invoice_ids']
        grand_taxable += g['taxable']
        grand_cgst    += g['cgst']
        grand_sgst    += g['sgst']
        grand_igst    += g['igst']

        row_num = sheet.max_row + 1
        sheet.append([
            f"{float(rate):.0f}%",     # e.g. "18%"  — plain string, no number_format
            len(g['invoice_ids']),
            _inr(g['taxable']),
            _inr(g['cgst']),
            _inr(g['sgst']),
            _inr(g['igst']),
            _inr(total_gst),
        ])
        _style_data_row(sheet, row_num, num_cols)

    # Grand-total row
    sheet.append([])
    totals_row = sheet.max_row + 1
    grand_gst = grand_cgst + grand_sgst + grand_igst
    sheet.append([
        'TOTAL',
        len(grand_inv_ids),
        _inr(grand_taxable),
        _inr(grand_cgst),
        _inr(grand_sgst),
        _inr(grand_igst),
        _inr(grand_gst),
    ])
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=totals_row, column=col)
        cell.font      = TOTALS_FONT
        cell.fill      = TOTALS_FILL
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
    sheet.row_dimensions[totals_row].height = 20

    sheet.append([])
    _merge_write(
        sheet, row=sheet.max_row + 1, start_col=1, end_col=num_cols,
        value='This report was generated automatically by EnterGYM.',
        font=FOOTER_FONT,
        alignment=Alignment(horizontal='left'),
    )
    _merge_write(
        sheet, row=sheet.max_row + 1, start_col=1, end_col=num_cols,
        value='For GST return filing, please consult your Chartered Accountant.',
        font=FOOTER_FONT,
        alignment=Alignment(horizontal='left'),
    )